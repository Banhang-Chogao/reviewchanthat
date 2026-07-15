#!/usr/bin/env python3
"""
Optional openpyxl generator for Movie_History_Template.xlsx
(Client-side SheetJS also generates the same structure in the browser.)

Usage:
  python3 scripts/movie_history_excel_template.py [output_path]
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as e:
    raise SystemExit(
        "openpyxl is required: pip install openpyxl\n" + str(e)
    )

HEADERS = [
    "Movie Title *",
    "Original Title",
    "Watch Date *",
    "Watch Time",
    "Country",
    "Language",
    "Genre",
    "Director",
    "Main Cast",
    "Streaming Platform",
    "Cinema",
    "Watching Method",
    "Duration (minutes)",
    "Personal Rating (0-10)",
    "IMDb Rating",
    "Rotten Tomatoes",
    "Letterboxd Rating",
    "Favorite (Yes/No)",
    "Rewatch (Yes/No)",
    "Watch With",
    "Mood",
    "Review",
    "Tags",
    "Poster URL (optional)",
    "Trailer URL",
    "Official Website",
    "Notes",
]

SAMPLE = [
    "THE ODYSSEY",
    "The Odyssey",
    "2026-07-17",
    "09:20",
    "USA",
    "English",
    "Action, Adventure",
    "Christopher Nolan",
    "Matt Damon, Tom Holland",
    "",
    "CGV Hùng Vương Plaza",
    "Cinema",
    150,
    9,
    8.2,
    "92%",
    4.1,
    "Yes",
    "No",
    "Friends",
    "Excited",
    "Epic Nolan adaptation — sample row for template only",
    "cinema, imax",
    "",
    "",
    "https://www.cgv.vn/",
    "Template sample",
]

INSTRUCTION = (
    "INSTRUCTIONS: Do not rename headers. Required: Movie Title *, Watch Date * "
    "(yyyy-mm-dd). Personal Rating 0–10. Favorite/Rewatch = Yes or No. "
    "Tags comma-separated. UTF-8 Vietnamese OK. One sample row below — replace or delete."
)

SDNA_FILL = PatternFill("solid", fgColor="00A7A0")
SDNA_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
INSTR_FILL = PatternFill("solid", fgColor="EEF8F7")
INSTR_FONT = Font(italic=True, color="555555", name="Calibri", size=10)
THIN = Border(
    left=Side(style="thin", color="007F7A"),
    right=Side(style="thin", color="007F7A"),
    top=Side(style="thin", color="007F7A"),
    bottom=Side(style="thin", color="007F7A"),
)


def build(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Movie History"

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.fill = SDNA_FILL
        cell.font = SDNA_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN

    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(2, col, INSTRUCTION if col == 1 else "")
        cell.fill = INSTR_FILL
        cell.font = INSTR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for col, val in enumerate(SAMPLE, 1):
        ws.cell(3, col, val)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 36

    for col, h in enumerate(HEADERS, 1):
        width = min(max(len(str(h)) + 2, 12), 36)
        sample_w = len(str(SAMPLE[col - 1])) + 2 if col <= len(SAMPLE) else 12
        ws.column_dimensions[get_column_letter(col)].width = min(max(width, sample_w), 36)

    # Yes/No validation for Favorite / Rewatch
    fav_col = HEADERS.index("Favorite (Yes/No)") + 1
    rew_col = HEADERS.index("Rewatch (Yes/No)") + 1
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(fav_col)}3:{get_column_letter(fav_col)}10000")
    dv.add(f"{get_column_letter(rew_col)}3:{get_column_letter(rew_col)}10000")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path}")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Movie_History_Template.xlsx")
    build(out)
