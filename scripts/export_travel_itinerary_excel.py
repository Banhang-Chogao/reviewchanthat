#!/usr/bin/env python3
"""
Export travel itinerary to Excel format.
Requires: openpyxl

Usage:
  python3 export_travel_itinerary_excel.py --trip '{"destination":"ICN","departure":"2026-10-15","...}'

Format:
  Multi-sheet workbook with Summary, Daily Plan, Budget, Packing, Transportation.
"""

import argparse
import json
import sys
from datetime import datetime, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


class TravelItineraryExcel:
    """Export travel itinerary to Excel."""

    def __init__(self, trip_data):
        self.trip = trip_data
        self.itinerary = trip_data.get('itinerary', {})
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

        # Styles
        self.header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF', size=12)
        self.subheader_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        self.subheader_font = Font(bold=True, size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def add_summary_sheet(self):
        """Add summary sheet."""
        ws = self.wb.create_sheet('Summary', 0)

        # Title
        ws['A1'] = 'Travel Itinerary Summary'
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:B1')

        # Data
        data = [
            ['Destination', self.trip.get('destination', 'N/A')],
            ['Departure Date', self.trip.get('departure', 'N/A')],
            ['Return Date', self.trip.get('returnDate', 'N/A')],
            ['Duration', f"{self.trip.get('days', 0)} days"],
            ['Number of Adults', self.trip.get('adults', 1)],
            ['Number of Children', self.trip.get('children', 0)],
            ['Budget Level', self.trip.get('budget', 'N/A')],
            ['Travel Purpose', ', '.join(self.trip.get('purposes', ['Tourism']))],
            ['Weather', self.itinerary.get('weather', 'N/A')],
            ['Hotel Area', self.itinerary.get('hotelArea', 'N/A')],
            ['Estimated Budget', self.itinerary.get('estimatedBudget', 'N/A')],
            ['Transportation', self.itinerary.get('transportation', 'N/A')],
            ['Weather Notes', self.itinerary.get('weatherNotes', 'N/A')],
            ['Local Tips', self.itinerary.get('localTips', 'N/A')],
            ['Emergency Numbers', self.itinerary.get('emergencyNumbers', 'N/A')],
            ['Airport Notes', self.itinerary.get('airportNotes', 'N/A')],
        ]

        for idx, (label, value) in enumerate(data, start=2):
            ws[f'A{idx}'] = label
            ws[f'A{idx}'].font = Font(bold=True)
            ws[f'B{idx}'] = str(value)
            ws[f'B{idx}'].alignment = Alignment(wrap_text=True)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

    def add_daily_plan_sheet(self):
        """Add daily plan sheet."""
        ws = self.wb.create_sheet('Daily Plan', 1)

        # Header
        headers = ['Day', 'Date', 'Morning', 'Lunch', 'Afternoon', 'Dinner', 'Notes']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(wrap_text=True)
            cell.border = self.border

        # Data
        start_date = datetime.strptime(self.trip.get('departure', '2026-01-01'), '%Y-%m-%d')
        daily = self.itinerary.get('dailyItinerary', [])

        for idx, day in enumerate(daily):
            date_obj = start_date + timedelta(days=idx)
            row = [
                f"Day {idx + 1}",
                date_obj.strftime('%Y-%m-%d'),
                day.get('morning', ''),
                day.get('lunch', ''),
                day.get('afternoon', ''),
                day.get('dinner', ''),
                day.get('notes', '')
            ]
            ws.append(row)

            for cell in ws[idx + 2]:
                cell.alignment = Alignment(wrap_text=True)
                cell.border = self.border

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 25

    def add_budget_sheet(self):
        """Add budget breakdown sheet."""
        ws = self.wb.create_sheet('Budget', 2)

        # Title
        ws['A1'] = 'Budget Breakdown'
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:C1')

        # Budget estimates based on level
        budget_level = self.trip.get('budget', 'mid-range')
        budget_ranges = {
            'luxury': {'daily': 300, 'hotel': 150, 'food': 80, 'activities': 70},
            'mid-range': {'daily': 150, 'hotel': 60, 'food': 50, 'activities': 40},
            'budget': {'daily': 80, 'hotel': 30, 'food': 30, 'activities': 20}
        }
        budget = budget_ranges.get(budget_level, budget_ranges['mid-range'])
        days = self.trip.get('days', 1)
        adults = self.trip.get('adults', 1)

        # Headers
        headers = ['Item', 'Daily', 'Total']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill

        # Data
        data = [
            ['Accommodation', budget['hotel'], budget['hotel'] * days],
            ['Meals', budget['food'], budget['food'] * days],
            ['Activities', budget['activities'], budget['activities'] * days],
            ['Miscellaneous', budget['daily'] - budget['hotel'] - budget['food'] - budget['activities'],
             (budget['daily'] - budget['hotel'] - budget['food'] - budget['activities']) * days],
        ]

        for row_idx, row_data in enumerate(data, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx == 1:
                    cell.value = value
                    cell.font = Font(bold=True)
                else:
                    cell.value = value
                    cell.number_format = '$#,##0.00'
                cell.border = self.border

        # Total
        total_row = len(data) + 4
        ws.cell(row=total_row, column=1).value = 'TOTAL'
        ws.cell(row=total_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=total_row, column=2).value = budget['daily']
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        ws.cell(row=total_row, column=2).number_format = '$#,##0.00'
        ws.cell(row=total_row, column=3).value = budget['daily'] * days * adults
        ws.cell(row=total_row, column=3).font = Font(bold=True, size=12, color='FF0000')
        ws.cell(row=total_row, column=3).number_format = '$#,##0.00'

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def add_packing_sheet(self):
        """Add packing list sheet."""
        ws = self.wb.create_sheet('Packing', 3)

        # Title
        ws['A1'] = 'Packing List'
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:B1')

        # Headers
        ws['A2'] = 'Item'
        ws['B2'] = 'Packed'
        for cell in [ws['A2'], ws['B2']]:
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.border = self.border

        # Packing list
        packing = self.itinerary.get('packingList', [])
        for idx, item in enumerate(packing, start=3):
            ws[f'A{idx}'] = item
            ws[f'B{idx}'] = '☐'
            ws[f'A{idx}'].border = self.border
            ws[f'B{idx}'].border = self.border
            ws[f'B{idx}'].alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 8

    def add_transportation_sheet(self):
        """Add transportation info sheet."""
        ws = self.wb.create_sheet('Transportation', 4)

        # Title
        ws['A1'] = 'Transportation Information'
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:B1')

        # Content
        sections = [
            ('Getting Around', self.itinerary.get('transportation', 'N/A')),
            ('Airport Arrival', self.itinerary.get('airportNotes', 'N/A')),
            ('Local Tips', self.itinerary.get('localTips', 'N/A')),
        ]

        row = 3
        for title, content in sections:
            ws[f'A{row}'] = title
            ws[f'A{row}'].font = self.subheader_font
            row += 1

            ws[f'A{row}'] = str(content)
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'A{row}:B{row}')
            row += 2

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 50

    def generate(self):
        """Generate complete workbook."""
        self.add_summary_sheet()
        self.add_daily_plan_sheet()
        self.add_budget_sheet()
        self.add_packing_sheet()
        self.add_transportation_sheet()
        return self.wb


def main():
    parser = argparse.ArgumentParser(description='Export travel itinerary to Excel')
    parser.add_argument('--trip', required=True, help='Trip data as JSON string')
    parser.add_argument('--output', default=None, help='Output file path (optional)')
    args = parser.parse_args()

    try:
        trip_data = json.loads(args.trip)
    except json.JSONDecodeError as e:
        print(f'Invalid JSON: {e}')
        sys.exit(1)

    excel = TravelItineraryExcel(trip_data)
    wb = excel.generate()

    if args.output:
        wb.save(args.output)
        print(f'Excel saved to {args.output}')
    else:
        import tempfile
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            with open(tmp.name, 'rb') as f:
                sys.stdout.buffer.write(f.read())


if __name__ == '__main__':
    main()
